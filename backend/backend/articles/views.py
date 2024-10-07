from rest_framework import generics
from rest_framework import status
from rest_framework.views import APIView
from rest_framework.response import Response
from django.http import HttpResponse

from .models import Article
from .serializers import ArticleSerializer

from openpyxl import load_workbook, Workbook

print("ExcelView test")


class ExcelView(APIView):

    def get(self, request, *args, **kwargs):

        print(self, request)

        workbook = Workbook()

        sheet = workbook.active

        sheet.title = 'articles'

        headers = [field.name for field in Article._meta.fields]

        sheet.append(headers)

        queryset = Article.objects.all()

        serializer = ArticleSerializer(queryset, many=True)

        for data in serializer.data:
            sheet.append([data[header] for header in headers])

        # Crear una respuesta HTTP para enviar el archivo Excel
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        response['Content-Disposition'] = 'attachment; filename=datos.xlsx'
        
        workbook.save(response)

        return response

    def post(self, request, *args, **kwargs):

        if 'file' not in request.FILES:
            return Response({'error': 'No se ha subido ningún archivo.'}, status=status.HTTP_400_BAD_REQUEST)

        excel_file = request.FILES['file']

        try:
            workbook = load_workbook(excel_file, read_only=True)
            sheet = workbook.active

            rows = list(sheet.iter_rows(values_only=True))
            headers = rows[0]
            data = []

            for row in rows[1:]:
                row_data = dict(zip(headers, row))
                data.append(row_data)

            for item in data:
                existing_article = Article.objects.filter(code=item['code']).first()
                
                if existing_article:
                    serializer = ArticleSerializer(existing_article, data=item)
                    if serializer.is_valid():
                        serializer.save()
                    else:
                        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
                else:
                    # Si el artículo no existe, lo crea
                    serializer = ArticleSerializer(data=item)
                    if serializer.is_valid():
                        serializer.save()
                    else:
                        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

            return Response({'message': 'Datos importados!'}, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class ArticleListCreate(generics.ListCreateAPIView):
    queryset = Article.objects.all()
    serializer_class = ArticleSerializer


class ArticleDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = Article.objects.all()
    serializer_class = ArticleSerializer
